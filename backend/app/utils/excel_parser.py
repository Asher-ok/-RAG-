"""
Excel/CSV文件解析工具
"""
import io
import pandas as pd
from typing import List, Dict, Any
from openpyxl import load_workbook


class ExcelParser:
    """Excel/CSV解析器"""
    
    @staticmethod
    def parse_excel(file_content: bytes, file_type: str = "xlsx") -> Dict[str, Any]:
        """
        解析Excel/CSV文件
        
        Args:
            file_content: 文件二进制内容
            file_type: 文件类型 xlsx/csv
            
        Returns:
            解析结果字典
        """
        try:
            if file_type == "csv":
                df = pd.read_csv(io.BytesIO(file_content), encoding='utf-8')
            else:
                # 尝试读取Excel，跳过可能的说明行
                try:
                    # 先尝试正常读取
                    df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
                except Exception:
                    # 如果失败，尝试跳过第一行（说明行）
                    df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', header=1)
                
                # 检查是否第一行是说明行（包含"填写说明"等关键字）
                if '填写说明' in str(df.columns[0]) or '填写说明：' in str(df.iloc[0, 0] if len(df) > 0 else ''):
                    # 重新读取，跳过说明行
                    df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl', header=1)
            
            # 验证必需列
            required_columns = ['商品标题', '一级类目', '二级类目', '三级类目', '图片URL', 'SKU列表']
            # 兼容抖店官方导出的字段名
            if '商品名称' in df.columns:
                df.rename(columns={'商品名称': '商品标题'}, inplace=True)
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return {
                    "success": False,
                    "message": f"缺少必需列: {', '.join(missing_columns)}"
                }
            
            # 解析商品数据
            products = []
            invalid_rows = []
            
            for index, row in df.iterrows():
                try:
                    product = ExcelParser._parse_product_row(row, index)
                    if product:
                        products.append(product)
                except Exception as e:
                    invalid_rows.append({
                        "row": index + 3,  # Excel行号从1开始，加上表头和说明行
                        "error": str(e)
                    })
            
            return {
                "success": True,
                "products": products,
                "total_count": len(products),
                "invalid_rows": invalid_rows
            }
            
        except Exception as e:
            return {
                "success": False,
                "message": f"文件解析失败: {str(e)}"
            }
    
    @staticmethod
    def _parse_product_row(row: pd.Series, index: int) -> Dict[str, Any]:
        """解析单行商品数据（完全兼容抖店官方字段）"""
        # 解析图片URL（支持逗号或分号分隔）
        images_str = str(row.get('图片URL', '')).strip()
        images = [img.strip() for img in images_str.replace('；', ';').split(';') if img.strip()]
        if not images:
            images = [img.strip() for img in images_str.split(',') if img.strip()]
        
        # 解析SKU列表（格式：SKU_ID:SKU名称:价格:库存;...）
        sku_list_str = str(row.get('SKU列表', '')).strip()
        sku_list = ExcelParser._parse_sku_list(sku_list_str, row)
        
        if not sku_list:
            raise ValueError(f"第{index + 2}行：SKU列表为空或格式错误")
        
        product = {
            # 基础信息
            "title": str(row.get('商品名称', row.get('商品标题', ''))).strip(),
            
            # 类目信息
            "first_cid": str(row.get('一级类目', '')).strip(),
            "second_cid": str(row.get('二级类目', '')).strip(),
            "third_cid": str(row.get('三级类目', '')).strip(),
            "fourth_cid": str(row.get('四级类目', '')).strip() if pd.notna(row.get('四级类目')) else None,
            
            # 商品类型和分组
            "product_type": int(row.get('商品类型', 1)) if pd.notna(row.get('商品类型')) else 1,
            "product_group": str(row.get('商品分组', '')).strip() if pd.notna(row.get('商品分组')) else None,
            
            # 商家编码
            "merchant_code": str(row.get('商家编码', '')).strip() if pd.notna(row.get('商家编码')) else None,
            "item_number": str(row.get('货号', '')).strip() if pd.notna(row.get('货号')) else None,
            
            # 图片和SKU
            "images": images,
            "sku_list": sku_list,
            
            # 发货和销售
            "delivery_time": str(row.get('商品发货时间', '')).strip() if pd.notna(row.get('商品发货时间')) else None,
            "sales_count": int(row.get('销量', 0)) if pd.notna(row.get('销量')) else 0,
            
            # 佣金和审核
            "commission_rate": float(row.get('佣金比例', 0)) if pd.notna(row.get('佣金比例')) else None,
            "audit_status": int(row.get('商品审核状态', 0)) if pd.notna(row.get('商品审核状态')) else 0,
            
            # 链接
            "product_url": str(row.get('商品链接', '')).strip() if pd.notna(row.get('商品链接')) else None,
        }
        
        # 验证必填字段
        if not product['title']:
            raise ValueError(f"第{index + 2}行：商品名称不能为空")
        if not product['first_cid'] or not product['second_cid'] or not product['third_cid']:
            raise ValueError(f"第{index + 2}行：类目信息不完整")
        if not product['images']:
            raise ValueError(f"第{index + 2}行：图片URL不能为空")
        
        return product
    
    @staticmethod
    def _parse_sku_list(sku_str: str, row: pd.Series = None) -> List[Dict[str, Any]]:
        """
        解析SKU列表字符串（完全兼容抖店官方字段）
        格式：SKU_ID:SKU名称:价格:库存:商家SKU编码:规格ID:规格名称:现货可售:预售库存
        """
        if not sku_str or sku_str == 'nan':
            return []
        
        sku_list = []
        # 支持分号或换行分隔
        sku_items = sku_str.replace('\n', ';').split(';')
        
        for sku_item in sku_items:
            sku_item = sku_item.strip()
            if not sku_item:
                continue
            
            # 分割SKU信息（支持冒号或逗号分隔）
            parts = sku_item.replace('，', ',').split(':')
            if len(parts) < 4:
                parts = sku_item.split(',')
            
            if len(parts) >= 4:
                try:
                    sku = {
                        "sku_id": parts[0].strip(),
                        "sku_name": parts[1].strip(),
                        "price": int(float(parts[2].strip()) * 100),  # 元转分
                        "stock": int(parts[3].strip()),
                        "merchant_sku_code": parts[4].strip() if len(parts) > 4 else None,
                        "spec_id": parts[5].strip() if len(parts) > 5 else None,
                        "spec_name": parts[6].strip() if len(parts) > 6 else None,
                        "available_stock": int(parts[7].strip()) if len(parts) > 7 and parts[7].strip() else None,
                        "presale_stock": int(parts[8].strip()) if len(parts) > 8 and parts[8].strip() else None,
                    }
                    sku_list.append(sku)
                except (ValueError, IndexError):
                    continue
        
        # 如果从row中能获取到商家SKU编码，也添加进去
        if row is not None and pd.notna(row.get('商家SKU编码')):
            merchant_sku_code = str(row.get('商家SKU编码', '')).strip()
            if merchant_sku_code and len(sku_list) > 0:
                sku_list[0]['merchant_sku_code'] = merchant_sku_code
        
        return sku_list
    
    @staticmethod
    def generate_template() -> bytes:
        """
        生成Excel模板文件（完全兼容抖店官方字段）
        
        Returns:
            Excel文件二进制内容
        """
        # 创建示例数据
        data = {
            '商品名称': [
                '【示例】时尚女装连衣裙夏季新款',
                '【示例】男士休闲运动鞋透气舒适'
            ],
            '一级类目': ['20100', '20100'],
            '二级类目': ['20101', '20101'],
            '三级类目': ['20102', '20102'],
            '四级类目': ['', ''],
            '商品类型': [1, 1],
            '商品分组': ['女装', '男鞋'],
            '商家编码': ['MC001', 'MC002'],
            '货号': ['HN2024001', 'HN2024002'],
            '图片URL': [
                'https://example.com/image1.jpg;https://example.com/image2.jpg;https://example.com/image3.jpg',
                'https://example.com/image4.jpg;https://example.com/image5.jpg'
            ],
            'SKU列表': [
                'SKU001:红色-S码:99.00:100:MSKU001:SPEC001:颜色规格:80:20;SKU002:红色-M码:99.00:80:MSKU002:SPEC001:颜色规格:60:20',
                'SKU004:黑色-39码:199.00:200:MSKU004:SPEC002:尺码规格:180:20;SKU005:白色-40码:199.00:150:MSKU005:SPEC002:尺码规格:130:20'
            ],
            '商品发货时间': ['24小时内发货', '48小时内发货'],
            '佣金比例': [10, 15],
            '商品链接': ['', '']
        }
        
        df = pd.DataFrame(data)
        
        # 写入Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='商品数据')
            
            # 获取工作表并设置列宽和说明
            worksheet = writer.sheets['商品数据']
            worksheet.column_dimensions['A'].width = 35  # 商品名称
            worksheet.column_dimensions['B'].width = 12  # 一级类目
            worksheet.column_dimensions['C'].width = 12  # 二级类目
            worksheet.column_dimensions['D'].width = 12  # 三级类目
            worksheet.column_dimensions['E'].width = 12  # 四级类目
            worksheet.column_dimensions['F'].width = 12  # 商品类型
            worksheet.column_dimensions['G'].width = 15  # 商品分组
            worksheet.column_dimensions['H'].width = 15  # 商家编码
            worksheet.column_dimensions['I'].width = 15  # 货号
            worksheet.column_dimensions['J'].width = 60  # 图片URL
            worksheet.column_dimensions['K'].width = 100 # SKU列表
            worksheet.column_dimensions['L'].width = 20  # 商品发货时间
            worksheet.column_dimensions['M'].width = 12  # 佣金比例
            worksheet.column_dimensions['N'].width = 50  # 商品链接
            
            # 添加说明行（在第一行上方插入）
            worksheet.insert_rows(1)
            worksheet['A1'] = '填写说明：必填字段不能为空'
            worksheet['B1'] = '从抖店后台获取'
            worksheet['E1'] = '可选'
            worksheet['F1'] = '1=普通/2=虚拟'
            worksheet['G1'] = '可选'
            worksheet['H1'] = '可选，内部编码'
            worksheet['I1'] = '可选'
            worksheet['J1'] = '多个用分号(;)分隔'
            worksheet['K1'] = 'SKU编码:名称:价格:库存:商家SKU编码:规格ID:规格名称:现货:预售（后5项可选）'
            worksheet['L1'] = '可选，如"24小时内"'
            worksheet['M1'] = '可选，如10表示10%'
            worksheet['N1'] = '可选'
            
            # 设置说明行样式
            from openpyxl.styles import Font, PatternFill
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color='FF0000', size=9)
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        output.seek(0)
        return output.read()
